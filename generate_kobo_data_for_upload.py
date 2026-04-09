"""
Generate Tool Rental Data for Manual Upload to Kobo Toolbox
Since the API upload requires special permissions, this script generates
properly formatted files that can be uploaded via the Kobo web interface.
"""

import json
import csv
from datetime import datetime, timedelta
import random

# Sample data
TOOLS = [
    "Hammer", "Drill", "Saw", "Screwdriver Set", "Wrench Set",
    "Pliers", "Level", "Tape Measure", "Ladder", "Power Drill",
    "Circular Saw", "Sander", "Paint Roller", "Garden Hose", "Wheelbarrow"
]

BORROWERS = [
    "Jesse Artache", "Isaac Simkin", "Israel Simmons", "Lydia Anwor",
    "Danek Obriondo", "Derek Obriondo", "Clement Jackson", "Cassanova Miller",
    "Maria Rodriguez", "John Thompson", "Sarah Williams", "Michael Chen", "Patricia Davis"
]

CONDITIONS = [
    "Good condition upon return", "Good Condition upon Return",
    "Excellent condition", "Minor wear, acceptable", "Needs cleaning", "Small scratch noted"
]

DAMAGE_CHARGES = ["—", "—", "—", "—", "$5 cleaning fee", "$10 minor damage", "$15 repair needed", "No charge"]

def generate_sample_data(num_records=30):
    """Generate sample tool rental records"""
    records = []
    start_date = datetime.now() - timedelta(days=90)
    
    for i in range(num_records):
        loan_date = start_date + timedelta(days=random.randint(0, 90))
        loan_time = f"{random.randint(8, 17):02d}:{random.choice(['00', '15', '30', '45'])}"
        
        days_borrowed = random.randint(1, 14)
        return_date = loan_date + timedelta(days=days_borrowed)
        return_time = f"{random.randint(8, 17):02d}:{random.choice(['00', '15', '30', '45'])}"
        
        record = {
            "date_loaned": loan_date.strftime("%Y-%m-%d"),
            "time_loaned": loan_time,
            "date_returned": return_date.strftime("%Y-%m-%d"),
            "time_returned": return_time,
            "tool_name": random.choice(TOOLS),
            "borrower_name": random.choice(BORROWERS),
            "borrower_signature": random.choice(BORROWERS).split()[0],
            "condition_upon_return": random.choice(CONDITIONS),
            "damage_charged": random.choice(DAMAGE_CHARGES),
            "return_notes": random.choice([
                "On time return", "Late return, no charge", "Returned clean",
                "Needs inspection", "Good borrower", ""
            ]),
            "serial_number": f"TOOL-{1000 + i}",
            "quantity": random.randint(1, 4)
        }
        
        records.append(record)
    
    return records

def create_excel_file(records, filename="tool_rental_kobo_import.xlsx"):
    """Create Excel file for Kobo import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Header row with formatting
        headers = list(records[0].keys()) if records else []
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=record[key])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"✓ Excel file created: {filename}")
        return True
        
    except Exception as e:
        print(f"✗ Error creating Excel file: {e}")
        return False

def save_json(records, filename="tool_rental_kobo_import.json"):
    """Save as JSON"""
    with open(filename, 'w') as f:
        json.dump(records, f, indent=2)
    print(f"✓ JSON file created: {filename}")

def save_csv(records, filename="tool_rental_kobo_import.csv"):
    """Save as CSV"""
    if not records:
        return
    
    with open(filename, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)
    print(f"✓ CSV file created: {filename}")

def print_upload_instructions():
    """Print instructions for manual upload"""
    print("\n" + "=" * 70)
    print("MANUAL UPLOAD INSTRUCTIONS")
    print("=" * 70)
    print("\nTo upload this data to your Kobo form:")
    print("\n1. Go to your form in Kobo Toolbox:")
    print("   https://kf.kobotoolbox.org/#/forms/aJ7GEDZPU3dbM4KAEqUBKW/")
    print("\n2. Click on the 'DATA' tab")
    print("\n3. Click the 'Upload' button (or import icon)")
    print("\n4. Select one of the generated files:")
    print("   - tool_rental_kobo_import.xlsx (RECOMMENDED)")
    print("   - tool_rental_kobo_import.csv")
    print("\n5. Follow the on-screen prompts to complete the upload")
    print("\n6. Your 30 sample records will be imported!")
    print("\n" + "=" * 70)

def main():
    print("=" * 70)
    print("Tool Rental Data Generator for Kobo Toolbox")
    print("=" * 70)
    
    # Generate data
    print("\nGenerating 30 sample tool rental records...")
    records = generate_sample_data(30)
    print(f"✓ Generated {len(records)} records")
    
    # Show sample
    print("\nSample record:")
    print("-" * 70)
    for key, value in list(records[0].items())[:6]:
        print(f"  {key:25s}: {value}")
    print("  ...")
    
    # Create files
    print("\nCreating export files...")
    print("-" * 70)
    save_json(records)
    save_csv(records)
    create_excel_file(records)
    
    # Print instructions
    print_upload_instructions()

if __name__ == "__main__":
    main()
